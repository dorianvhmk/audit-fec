import io
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import PatternFill, Font

from app.services.supabase_store import (
    get_analysis,
    list_analyses,
    get_analysis_step,
    STEP_LABELS,
)

router = APIRouter(tags=["results"])

_STATUS_COLORS = {
    "OK":     "C6EFCE",  # green
    "écart":  "FFEB9C",  # yellow
    "erreur": "FFC7CE",  # red
    "absent": "E0E0E0",  # gray
}


@router.get("/progress/{analysis_id}")
def get_progress(analysis_id: str):
    """
    Return the current pipeline step for an in-flight analysis.

    Response shape::

        {
          "status":          "pending" | "processing" | "done" | "error",
          "step":            "parsing_fec" | "extracting_pdf" | "reconciling"
                             | "generating_comments" | "done" | "",
          "step_label":      str,        # human-readable French label
          "steps_completed": int,        # 0-based count of finished steps
          "steps_total":     int,        # total number of steps (4)
        }
    """
    record = get_analysis(analysis_id)
    if not record:
        raise HTTPException(404, "Analyse introuvable")

    status = record["status"]
    step_keys = list(STEP_LABELS.keys())
    n_steps = len(step_keys)

    if status == "done":
        return {
            "status": "done",
            "step": "done",
            "step_label": "Terminé",
            "steps_completed": n_steps,
            "steps_total": n_steps,
        }

    if status == "error":
        step = get_analysis_step(analysis_id) or ""
        idx = step_keys.index(step) if step in step_keys else 0
        return {
            "status": "error",
            "step": step,
            "step_label": STEP_LABELS.get(step, "Erreur"),
            "steps_completed": idx,
            "steps_total": n_steps,
        }

    # pending or processing
    step = get_analysis_step(analysis_id) or ""
    idx = (step_keys.index(step) + 1) if step in step_keys else 0
    return {
        "status": status,
        "step": step,
        "step_label": STEP_LABELS.get(step, "Initialisation…"),
        "steps_completed": idx,
        "steps_total": n_steps,
    }


@router.get("/analyses")
def get_analyses():
    """List the 50 most recent analyses (id, client_name, status, created_at)."""
    return list_analyses(limit=50)


@router.get("/results/{analysis_id}")
def get_results(analysis_id: str):
    record = get_analysis(analysis_id)
    if not record:
        raise HTTPException(404, "Analyse introuvable")
    return record


@router.get("/export/{analysis_id}")
def export_excel(analysis_id: str):
    record = get_analysis(analysis_id)
    if not record:
        raise HTTPException(404, "Analyse introuvable")
    if record["status"] != "done" or not record.get("results"):
        raise HTTPException(400, "Analyse non terminée")

    rows = record["results"].get("rows", [])
    client_name = record.get("client_name", "")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapprochement"

    headers = ["Poste", "Section", "Montant plaquette", "Montant FEC", "Écart (€)", "Écart (%)", "Statut", "Commentaire"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        status = row.get("status", "")
        ws.append([
            row.get("label"),
            row.get("section"),
            row.get("plaquette_amount"),
            row.get("fec_amount"),
            row.get("delta_abs"),
            row.get("delta_pct"),
            status,
            row.get("commentary"),
        ])
        fill_color = _STATUS_COLORS.get(status, "FFFFFF")
        fill = PatternFill(fill_type="solid", fgColor=fill_color)
        for cell in ws[ws.max_row]:
            cell.fill = fill

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["G"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = "".join(c for c in client_name if c.isalnum() or c in " _-")
    filename = f"rapprochement_{safe_name}.xlsx".replace(" ", "_")

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
